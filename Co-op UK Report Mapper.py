import io
import re
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from zoneinfo import ZoneInfo

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries


APP_TITLE = "Co-op UK Report Generator"
REPORT_PREFIX = "Co-op UK Report "
CLIENT_KEY = "coopuk"


NORMAL_HEADERS = [
    "Internal ID",
    "Site Name",
    "Site Address 1",
    "Site Address 2",
    "Site Post Code",
    "Item to Order",
    "Date of Visit",
    "Time of Visit",
    "Site Code",
    "Primary Result",
    "Please detail why you were unable to conduct this audit:",
    "What was the Operator number from the receipt?",
    "From the top of the receipt, please enter the store name and any visible codes.",
    "What type of alcohol did you try to purchase?",
    "Please give details of the product that you tried to purchase:",
    "How many people were in the queue?",
    "At which type of till was the purchase made?",
    "Did you make the purchase on its own or as part of a larger shop?",
    "Was there any generic 'Think 25' or 'Think 21' material visible from the till?",
    "Was the staff member who served you working entirely alone?",
    "Did the staff member who served you make eye contact with you during the transaction?",
    "When was eye contact first made?",
    "Did the staff member who served you look at you long enough to assess your age?",
    "Did the staff member who served you ask your age?",
    "Was the cabinet open before the staff member retrieved your cigarettes?",
    "Did the staff member who served you ask for ID?",
    "Were you asked for ID before the cigarettes had been retrieved from the cabinet?",
    "If cigarettes were retrieved, was the cabinet shut immediately after the staff member had retrieved your cigarettes?",
    "Was a supervisor called at any point during the transaction?",
    "Please use this space to explain anything unusual about your visit or to clarify any detail of your report:",
]


RAPID_HEADERS = [
    "Internal ID",
    "Site Name",
    "Site Address 1",
    "Site Address 2",
    "Site Post Code",
    "Item to Order",
    "Date of Visit",
    "Time of Visit",
    "Site Code",
    "Primary Result",
    "Please detail why you were unable to conduct this audit:",
    "Please enter the order number from your online receipt:",
    "Please enter the date you placed your order:",
    "Please enter the time you placed your order:",
    "Please select the service provider used:",
    "Please detail the store name you ordered it from:",
    "Please confirm the postcode of the store you ordered it from:",
    "What is your age?",
    "What was the total cost of your purchase?",
    "Please give details of the age restricted product(s) purchased:",
    "Did you order via the app or website?",
    "Please state the Co-op store name and address you ordered from:",
    "Was it easy to find the store on the app or website?",
    "Was the app or website easy to use?",
    "Was your order delivered within 30 minutes of placing the order?",
    "Did the driver ask your age?",
    "Did the driver ask you to enter your date of birth into a device?",
    "Did the driver ask for ID?",
    "Were any of the items damaged?",
    "Were any of the items missing?",
    "Please confirm what items were actually missing:",
    "Did the delivery bag arrive sealed with a sticker / sealed by staples / bag was not sealed?",
    "Did the delivery bag have a Think 25 or age restricted type sticker?",
    "Was the delivery driver dressed in branded attire coinciding with the app/website used?",
    "Did the driver make eye contact with you during the interaction?",
    "Was the driver friendly?",
    "Based on your online shopping experience, please rate the service from 1 to 10 (where 1 is very poor and 10 is excellent):",
    "Please explain the reason for your score:",
    "Based on your delivery experience, please rate your experience from 1 to 10 (where 1 is very poor and 10 is excellent):",
    "Were you able to order all items you wanted?",
    "Was the item you wanted to order unavailable?",
    "What was the item you were unable to order?",
    "Please use this space to explain anything unusual about your visit or to clarify any detail of your report:",
]


@dataclass(frozen=True)
class SourceSpec:
    columns: tuple[str, ...]
    combine: bool = False


def source(*columns, combine=False):
    return SourceSpec(tuple(columns), combine)


BASE_MAPPING = {
    "Internal ID": source("internal_id"),
    "Site Name": source("site_name"),
    "Site Address 1": source("site_address_1"),
    "Site Address 2": source("site_address_2"),
    "Site Post Code": source("site_post_code"),
    "Item to Order": source("item_to_order"),
    "Site Code": source("site_code"),
    "Primary Result": source("primary_result"),
    "Please detail why you were unable to conduct this audit:": source(
        "Please detail why you were unable to conduct this audit:"
    ),
}


NORMAL_MAPPING = {
    **BASE_MAPPING,
    "What was the Operator number from the receipt?": source(
        "What was the Operator number from the receipt?"
    ),
    "From the top of the receipt, please enter the store name and any visible codes.": source(
        "From the top of the receipt, please enter the store name and any visible codes.",
        "From the top of the receipt, please enter the store name and any visible codes:",
    ),
    "What type of alcohol did you try to purchase?": source(
        "What type of alcohol did you try to purchase?"
    ),
    "Please give details of the product that you tried to purchase:": source(
        "Please give details of the alcohol that you purchased:",
        "Please give details of the e-cig product that you tried to purchase:",
        "Please give details of the e-cig product that you purchased:",
        "Please give details of the cigarettes that you tried to purchase:",
        combine=True,
    ),
    "How many people were in the queue?": source(
        "How many people were in the queue?"
    ),
    "At which type of till was the purchase made?": source(
        "At which type of till was the purchase made?"
    ),
    "Did you make the purchase on its own or as part of a larger shop?": source(
        "Did you make the purchase on its own or as part of a larger shop?"
    ),
    "Was there any generic 'Think 25' or 'Think 21' material visible from the till?": source(
        "Was there any generic 'Think 25' or 'Think 21' material visible from the till?"
    ),
    "Was the staff member who served you working entirely alone?": source(
        "Was the staff member who served you working entirely alone?"
    ),
    "Did the staff member who served you make eye contact with you during the transaction?": source(
        "Did the staff member who served you make eye contact with you during the transaction?"
    ),
    "When was eye contact first made?": source("When was eye contact first made?"),
    "Did the staff member who served you look at you long enough to assess your age?": source(
        "Did the staff member who served you look at you long enough to assess your age?",
        "Did the staff member who served you look at you long enough to assess your age?  ",
    ),
    "Did the staff member who served you ask your age?": source(
        "Did the staff member who served you ask your age?"
    ),
    "Was the cabinet open before the staff member retrieved your cigarettes?": source(
        "Was the cabinet open before the staff member retrieved your cigarettes?"
    ),
    "Did the staff member who served you ask for ID?": source(
        "Did the staff member who served you ask for ID?"
    ),
    "Were you asked for ID before the cigarettes had been retrieved from the cabinet?": source(
        "Were you asked for ID before the cigarettes had been retrieved from the cabinet?"
    ),
    "If cigarettes were retrieved, was the cabinet shut immediately after the staff member had retrieved your cigarettes?": source(
        "If cigarettes were retrieved, was the cabinet shut immediately after the staff member had retrieved your cigarettes?"
    ),
    "Was a supervisor called at any point during the transaction?": source(
        "Was a supervisor called at any point during the transaction?"
    ),
    "Please use this space to explain anything unusual about your visit or to clarify any detail of your report:": source(
        "Please use this space to explain anything unusual about your visit or to clarify any detail of your report:"
    ),
}


RAPID_MAPPING = {
    **BASE_MAPPING,
    **{
        header: source(header)
        for header in RAPID_HEADERS[11:]
    },
}


REQUIRED_EXPORT_COLUMNS = {
    "client_name",
    "internal_id",
    "site_name",
    "site_address_1",
    "site_address_2",
    "site_post_code",
    "item_to_order",
    "site_code",
    "primary_result",
}


NORMAL_COMMON_QUESTIONS = [
    NORMAL_HEADERS[index]
    for index in (10, 11, 12, 15, 16, 17, 18, 19, 20, 21, 22, 23, 25, 28, 29)
]
NORMAL_ITEM_QUESTIONS = {
    "alcohol": [
        "What type of alcohol did you try to purchase?",
        "Please give details of the product that you tried to purchase:",
    ],
    "e-cig": ["Please give details of the product that you tried to purchase:"],
    "cigarettes": [
        "Please give details of the product that you tried to purchase:",
        "Was the cabinet open before the staff member retrieved your cigarettes?",
        "Were you asked for ID before the cigarettes had been retrieved from the cabinet?",
        "If cigarettes were retrieved, was the cabinet shut immediately after the staff member had retrieved your cigarettes?",
    ],
}


def clean_text(value):
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def normalise_header(value):
    return re.sub(r"\s+", " ", clean_text(value))


def normalise_key(value):
    return re.sub(r"[^a-z0-9]+", "", clean_text(value).lower())


def most_recent_saturday(as_of_date):
    return as_of_date - timedelta(days=(as_of_date.weekday() - 5) % 7)


def parse_date_value(value):
    text = clean_text(value)
    if not text:
        return None
    parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.to_pydatetime().replace(tzinfo=None)


def parse_time_value(value):
    text = clean_text(value)
    if not text:
        return None
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            pass
    return None


def smart_number(value, integer_only=False):
    text = clean_text(value)
    if not text:
        return None
    if re.fullmatch(r"[+-]?\d+", text):
        return int(text)
    if not integer_only and re.fullmatch(r"[+-]?(?:\d+\.\d*|\.\d+)", text):
        return float(text)
    return text


def read_audit_export(file_bytes):
    try:
        export = pd.read_csv(
            io.BytesIO(file_bytes),
            dtype=str,
            keep_default_na=False,
            encoding="utf-8-sig",
        )
    except UnicodeDecodeError:
        export = pd.read_csv(
            io.BytesIO(file_bytes),
            dtype=str,
            keep_default_na=False,
            encoding="cp1252",
        )

    missing = sorted(REQUIRED_EXPORT_COLUMNS - set(export.columns))
    if "date_of_visit_local" not in export.columns and "date_of_visit" not in export.columns:
        missing.append("date_of_visit_local (or date_of_visit)")
    if "time_of_visit_local" not in export.columns and "time_of_visit" not in export.columns:
        missing.append("time_of_visit_local (or time_of_visit)")
    if missing:
        raise ValueError(
            "The audit export is missing required column(s): " + ", ".join(missing)
        )
    return export


def worksheet_headers(worksheet, count):
    return [normalise_header(worksheet.cell(1, column).value) for column in range(1, count + 1)]


def find_report_sheet(workbook, preferred_name, expected_headers):
    expected = [normalise_header(value) for value in expected_headers]
    candidates = []
    if preferred_name in workbook.sheetnames:
        candidates.append(workbook[preferred_name])
    candidates.extend(ws for ws in workbook.worksheets if ws not in candidates)

    for worksheet in candidates:
        if worksheet_headers(worksheet, len(expected)) == expected:
            return worksheet

    raise ValueError(
        f"The previous report does not contain the expected '{preferred_name}' "
        "worksheet layout in row 1. Please upload the latest Co-op UK report."
    )


def last_data_row(worksheet):
    for row_number in range(worksheet.max_row, 1, -1):
        if clean_text(worksheet.cell(row_number, 1).value):
            return row_number
    return 1


def existing_report_ids(*worksheets):
    ids = set()
    for worksheet in worksheets:
        final_row = last_data_row(worksheet)
        for row_number in range(2, final_row + 1):
            audit_id = clean_text(worksheet.cell(row_number, 1).value)
            if audit_id:
                ids.add(audit_id)
    return ids


def copy_row_format(worksheet, source_row, target_row, max_column):
    if source_row < 2:
        return
    for column in range(1, max_column + 1):
        source_cell = worksheet.cell(source_row, column)
        target_cell = worksheet.cell(target_row, column)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format

    source_dimension = worksheet.row_dimensions[source_row]
    target_dimension = worksheet.row_dimensions[target_row]
    target_dimension.height = source_dimension.height
    target_dimension.hidden = source_dimension.hidden
    target_dimension.outlineLevel = source_dimension.outlineLevel


def value_from_spec(row, spec):
    values = []
    for column in spec.columns:
        if column in row.index:
            value = clean_text(row.get(column, ""))
            if value and value not in values:
                values.append(value)
    if not values:
        return None
    return " | ".join(values) if spec.combine else values[0]


def mapped_value(row, output_header, mapping, rapid=False):
    header = normalise_header(output_header)
    if header == "Date of Visit":
        return row["_visit_date"]
    if header == "Time of Visit":
        return row["_visit_time"]
    if rapid and header == "Site Post Code":
        return "-"

    value = value_from_spec(row, mapping.get(header, source()))
    if header in {
        "Site Code",
        "What was the Operator number from the receipt?",
        "How many people were in the queue?",
        "Please enter the order number from your online receipt:",
        "Based on your online shopping experience, please rate the service from 1 to 10 (where 1 is very poor and 10 is excellent):",
        "Based on your delivery experience, please rate your experience from 1 to 10 (where 1 is very poor and 10 is excellent):",
    }:
        return smart_number(value, integer_only=True)
    if header == "What was the total cost of your purchase?":
        return smart_number(value)
    if header == "Please enter the date you placed your order:":
        return parse_date_value(value) or value
    if header == "Please enter the time you placed your order:":
        return parse_time_value(value) or value
    return value


def mapping_by_normalised_header(mapping):
    return {normalise_header(header): spec for header, spec in mapping.items()}


def missing_relevant_questions(export, normal_rows, rapid_rows):
    checks = []
    if not normal_rows.empty:
        checks.extend(NORMAL_COMMON_QUESTIONS)
        present_items = {clean_text(value).lower() for value in normal_rows["item_to_order"]}
        for item_name, questions in NORMAL_ITEM_QUESTIONS.items():
            if item_name in present_items:
                checks.extend(questions)
    if not rapid_rows.empty:
        checks.extend(RAPID_HEADERS[11:])

    missing = []
    for header in dict.fromkeys(checks):
        mapping = NORMAL_MAPPING if header in NORMAL_MAPPING else RAPID_MAPPING
        spec = mapping.get(header)
        if spec and not any(column in export.columns for column in spec.columns):
            missing.append(header)
    return missing


def prepare_rows(export, existing_ids, cutoff_date):
    coop = export[
        export["client_name"].map(normalise_key).eq(CLIENT_KEY)
    ].copy()
    if coop.empty:
        raise ValueError("The export does not contain any Co-op UK audits.")

    coop["internal_id"] = coop["internal_id"].map(clean_text)
    missing_ids = int(coop["internal_id"].eq("").sum())
    coop = coop[coop["internal_id"].ne("")].copy()

    date_source = "date_of_visit_local" if "date_of_visit_local" in coop.columns else "date_of_visit"
    time_source = "time_of_visit_local" if "time_of_visit_local" in coop.columns else "time_of_visit"

    def row_date(row):
        return parse_date_value(row.get(date_source, ""))

    def row_time(row):
        return parse_time_value(row.get(time_source, ""))

    coop["_visit_date"] = coop.apply(row_date, axis=1)
    bad_dates = coop.loc[coop["_visit_date"].isna(), "internal_id"].tolist()
    if bad_dates:
        raise ValueError(
            "These Co-op UK audits have an invalid visit date: " + ", ".join(bad_dates)
        )

    after_cutoff = int(
        coop["_visit_date"].map(lambda value: value.date() > cutoff_date).sum()
    )
    coop = coop[
        coop["_visit_date"].map(lambda value: value.date() <= cutoff_date)
    ].copy()

    coop["_visit_time"] = coop.apply(row_time, axis=1)
    invalid_times = []
    for _, row in coop.iterrows():
        raw_time = clean_text(row.get(time_source, ""))
        if raw_time and row["_visit_time"] is None:
            invalid_times.append(row["internal_id"])
    if invalid_times:
        raise ValueError(
            "These Co-op UK audits have an invalid visit time: "
            + ", ".join(invalid_times)
        )

    before_deduplication = len(coop)
    coop = coop.drop_duplicates(subset=["internal_id"], keep="last")
    duplicate_count = before_deduplication - len(coop)
    already_reported = int(coop["internal_id"].isin(existing_ids).sum())
    coop = coop[~coop["internal_id"].isin(existing_ids)].copy()

    coop["_sort_time"] = coop["_visit_time"].map(
        lambda value: (
            value.hour,
            value.minute,
            value.second,
        ) if isinstance(value, time) else (24, 0, 0)
    )
    coop = coop.sort_values(
        ["_visit_date", "_sort_time", "internal_id"], kind="stable"
    )

    is_rapid = coop["item_to_order"].map(normalise_key).eq("rapiddelivery")
    normal_rows = coop[~is_rapid].copy()
    rapid_rows = coop[is_rapid].copy()
    details = {
        "missing_ids": missing_ids,
        "after_cutoff": after_cutoff,
        "duplicates": duplicate_count,
        "already_reported": already_reported,
    }
    return normal_rows, rapid_rows, details


def extend_sheet_range(worksheet, final_row, max_column):
    if worksheet.auto_filter.ref:
        worksheet.auto_filter.ref = f"A1:{get_column_letter(max_column)}{final_row}"
    for table in worksheet.tables.values():
        min_col, min_row, max_col, _ = range_boundaries(table.ref)
        if min_row == 1 and min_col == 1 and max_col == max_column:
            table.ref = f"A1:{get_column_letter(max_column)}{final_row}"


def append_rows(worksheet, rows, expected_headers, mapping, rapid=False):
    if rows.empty:
        return 0

    normalised_mapping = mapping_by_normalised_header(mapping)
    source_row = last_data_row(worksheet)
    first_new_row = source_row + 1

    for offset, (_, source_data) in enumerate(rows.iterrows()):
        target_row = first_new_row + offset
        copy_row_format(worksheet, source_row, target_row, len(expected_headers))
        for column_number in range(1, len(expected_headers) + 1):
            workbook_header = worksheet.cell(1, column_number).value
            worksheet.cell(target_row, column_number).value = mapped_value(
                source_data,
                workbook_header,
                normalised_mapping,
                rapid=rapid,
            )

    final_row = first_new_row + len(rows) - 1
    extend_sheet_range(worksheet, final_row, len(expected_headers))
    return len(rows)


def generate_report(audit_bytes, previous_report_bytes, as_of_date=None):
    if as_of_date is None:
        as_of_date = datetime.now(ZoneInfo("Europe/London")).date()
    elif isinstance(as_of_date, datetime):
        as_of_date = as_of_date.date()
    if not isinstance(as_of_date, date):
        raise TypeError("as_of_date must be a date or datetime")

    export = read_audit_export(audit_bytes)
    try:
        workbook = load_workbook(
            io.BytesIO(previous_report_bytes),
            data_only=False,
            keep_links=True,
        )
    except Exception as exc:
        raise ValueError("The previous report is not a valid .xlsx workbook.") from exc

    normal_sheet = find_report_sheet(workbook, "Report", NORMAL_HEADERS)
    rapid_sheet = find_report_sheet(workbook, "Rapid Delivery", RAPID_HEADERS)
    cutoff_date = most_recent_saturday(as_of_date)
    existing_ids = existing_report_ids(normal_sheet, rapid_sheet)
    normal_rows, rapid_rows, skipped = prepare_rows(
        export, existing_ids, cutoff_date
    )
    missing_questions = missing_relevant_questions(export, normal_rows, rapid_rows)

    normal_count = append_rows(
        normal_sheet,
        normal_rows,
        NORMAL_HEADERS,
        NORMAL_MAPPING,
    )
    rapid_count = append_rows(
        rapid_sheet,
        rapid_rows,
        RAPID_HEADERS,
        RAPID_MAPPING,
        rapid=True,
    )

    output_name = f"{REPORT_PREFIX}{as_of_date:%Y-%m-%d}.xlsx"
    total_count = normal_count + rapid_count
    if total_count == 0:
        output_bytes = previous_report_bytes
    else:
        workbook.calculation.calcMode = "auto"
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        output = io.BytesIO()
        workbook.save(output)
        output_bytes = output.getvalue()

    return output_bytes, output_name, {
        "new_count": total_count,
        "normal_count": normal_count,
        "rapid_count": rapid_count,
        "cutoff_date": cutoff_date,
        "missing_questions": missing_questions,
        **skipped,
    }


def main():
    st.title(APP_TITLE)
    st.write(
        "Upload the Co-op UK audit export and the previous Co-op UK Excel "
        "report. The generator removes audits already present, adds the new "
        "eligible audits to the correct worksheet, and keeps the workbook's "
        "existing layout and formatting."
    )

    audit_file = st.file_uploader(
        "1. Upload audits_basic_data_export.csv", type=["csv"]
    )
    previous_report_file = st.file_uploader(
        "2. Upload the previous Co-op UK report", type=["xlsx"]
    )

    all_uploaded = audit_file is not None and previous_report_file is not None
    if st.button("Generate report", type="primary", disabled=not all_uploaded):
        try:
            report_bytes, output_name, details = generate_report(
                audit_file.getvalue(),
                previous_report_file.getvalue(),
            )
        except Exception as exc:
            st.error(str(exc))
        else:
            count = details["new_count"]
            audit_word = "audit" if count == 1 else "audits"
            st.success(
                f"Report generated with {count} new {audit_word}: "
                f"{details['normal_count']} standard and "
                f"{details['rapid_count']} Rapid Delivery. "
                f"Audit cut-off: {details['cutoff_date']:%d/%m/%Y}."
            )

            skipped_parts = []
            if details["already_reported"]:
                skipped_parts.append(
                    f"{details['already_reported']} already in the report"
                )
            if details["duplicates"]:
                skipped_parts.append(
                    f"{details['duplicates']} duplicate IDs in the export"
                )
            if details["after_cutoff"]:
                skipped_parts.append(
                    f"{details['after_cutoff']} after the Saturday cut-off"
                )
            if details["missing_ids"]:
                skipped_parts.append(
                    f"{details['missing_ids']} without an Internal ID"
                )
            if skipped_parts:
                st.info("Skipped: " + "; ".join(skipped_parts) + ".")

            if details["missing_questions"]:
                st.warning(
                    "The export did not contain these relevant question columns, "
                    "so their report cells were left blank: "
                    + "; ".join(details["missing_questions"])
                )

            st.download_button(
                "Download updated Co-op UK report",
                data=report_bytes,
                file_name=output_name,
                mime=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
            )


if __name__ == "__main__":
    main()
